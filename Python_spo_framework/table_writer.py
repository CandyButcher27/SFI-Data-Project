"""
table_writer.py

Writes structured table extraction results from JSON into an Excel file.
- Creates or ensures existence of two sheets:
  1. 'Eligibility+EU Tax' — detailed criteria per Use of Proceeds
  2. 'SDG' — summary of SDGs per Use of Proceeds
- Handles appending to existing sheets safely, avoiding overwrite issues.
"""

import pandas as pd
import os
from openpyxl import load_workbook

def writer_to_excel_table(answer: dict, EXCEL_FILE: str):
    """
    Writes structured data from a parsed JSON (`answer`) into an Excel file.

    Args:
        answer (dict): JSON output containing 'Use_of_Proceeds', SDGs, and Eligibility Criteria.
        EXCEL_FILE (str): Path to the Excel file where data should be written.

    Behavior:
        - Creates the Excel file if it does not exist.
        - Ensures sheets 'Eligibility+EU Tax' and 'SDG' exist.
        - Appends new data to the sheets without overwriting existing rows.
        - Automatically generates a new Framework ID (F001, F002, ...).
    """

    sheet_elig = "Eligibility+EU Tax"
    sheet_sdg = "SDG"
    
    eligibility_headers = [
    "Framework ID",
    "Use of Proceeds",
    "Eligiblity Criteria",
    "SPO Evaluation",
    "EU Taxonomy Alignment",
    "DNSH",
    "Minimum Safeguards",
    "EU Taxonomy and Economic Activities"
    ]

    sdg_headers = [
    "Framework ID",
    "Use of Proceeds",
    "SDG"
    ]

    # --- Ensure Excel file exists and if not then create it with the relevant columns ---
    if not os.path.exists(EXCEL_FILE):
        
        df_eligibility = pd.DataFrame(columns=eligibility_headers)
        df_sdg = pd.DataFrame(columns=sdg_headers)
        
        # Create the workbook
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
            df_eligibility.to_excel(writer, sheet_name="Eligiblity+EU Tax", index=False)
            df_sdg.to_excel(writer, sheet_name="SDG", index=False)


    # --- Ensure sheets exist ---
    wb = load_workbook(EXCEL_FILE)
    
    for sheet_name,headers in [(sheet_elig, eligibility_headers), (sheet_sdg, sdg_headers)]:
        
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
        else:
            ws = wb[sheet_name]
            if ws.max_row==0:
                ws.appead(headers)         
            
    wb.save(EXCEL_FILE)
    wb.close()

    # --- Build dataframes from JSON ---
    use_of_proceeds = answer.get("Use_of_Proceeds", [])
    eligibility_rows = []
    sdg_rows = []

    # Determine next Framework ID
    def get_next_framework_id(ws):
        if ws.max_row < 2:
            return "F001"
        last_id = ws.cell(row=ws.max_row, column=1).value
        if not last_id or not last_id.startswith("F"):
            return "F001"
        number = int(last_id[1:])
        return f"F{number + 1:03d}"

    wb = load_workbook(EXCEL_FILE)
    ws_elig = wb[sheet_elig]
    framework_id = get_next_framework_id(ws_elig)
    wb.close()

    for uop in use_of_proceeds:
        name = uop.get("Name", "")
        sdgs = uop.get("SDGs", [])
        elig_list = uop.get("Eligibility_Criteria", [])

        # SDG sheet data
        sdg_combined = ", ".join(sdgs) if sdgs else ""
        sdg_rows.append({
            "Framework ID": framework_id,
            "Use of Proceeds": name,
            "SDG": sdg_combined
        })

        # Eligibility+EU Tax data
        for e in elig_list:
            eligibility_rows.append({
                "Framework ID": framework_id,
                "Use of Proceeds": name,
                "Eligibility Criteria": e.get("Description", ""),
                "SPO Evaluation": e.get("SPO_Evaluation", ""),
                "EU Taxonomy": e.get("EU_Taxonomy_Alignment", ""),
                "DNSH": e.get("DNSH", ""),
                "Minimum Safeguards": e.get("Minimum_Safeguards", ""),
                "NACE Code": e.get("NACE_Code", ""),
                "EU Taxonomy Economic Activity": e.get("EU_Taxonomy_Economic_Activity", "")
            })

    df_elig = pd.DataFrame(eligibility_rows)
    df_sdg = pd.DataFrame(sdg_rows)

    # --- Append to Excel safely ---
    with pd.ExcelWriter(EXCEL_FILE, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
        # Eligibility+EU Tax
        startrow = pd.read_excel(EXCEL_FILE, sheet_name=sheet_elig).shape[0]
        df_elig.to_excel(writer, sheet_name=sheet_elig, index=False, header=False, startrow=startrow)

        # SDG
        startrow = pd.read_excel(EXCEL_FILE, sheet_name=sheet_sdg).shape[0]
        df_sdg.to_excel(writer, sheet_name=sheet_sdg, index=False, header=False, startrow=startrow)

    print(f"✅ Data written successfully to {EXCEL_FILE} (Framework ID: {framework_id})")
